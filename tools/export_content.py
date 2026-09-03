"""Export every word the kiosk displays into a client-review workbook.

Reads the build's own content model, so what the client reviews is exactly what
a visitor sees — retyping it by hand is how the two drift apart. Re-run this
after any copy change, or the spreadsheet in the client's inbox is describing a
kiosk that no longer exists.

    python3 tools/export_content.py

Writes 'Unlock AI Value - content for client review.xlsx' to the repo root.
Requires openpyxl and node (to evaluate build/js/data.js).
"""
import json, datetime, os, subprocess, tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "Unlock AI Value - content for client review.xlsx")

# data.js is an ES module of plain literals; node is the only honest way to read
# it, and it keeps this export in step with the build automatically.
_dump = os.path.join(tempfile.gettempdir(), "uav_dump.mjs")
with open(_dump, "w", encoding="utf-8") as fh:
    fh.write("import * as D from %r;\n" % os.path.join(ROOT, "build/js/data.js")
             + "const o={};for(const[k,v]of Object.entries(D))o[k]=typeof v==='function'?'[fn]':v;"
             # resolved here rather than reimplemented in Python, so the sheet
             # quotes the string the kiosk actually shows
               "o.BENCHMARK_FINE_PRINT=D.benchmarkFinePrint();"
               "process.stdout.write(JSON.stringify(o));")
D = json.loads(subprocess.check_output(["node", _dump], text=True))

NAME = {p["id"]: p["name"] for p in D["POOLS"]}
ORDER = [p["id"] for p in sorted(D["POOLS"], key=lambda p: p["edge"])]
BANDS = D["BANDS"]
RANGES = ["0-2", "3-5", "6-8", "9-10", "11-12"]
REVIEW = ["Approved? (Y/N)", "Revised copy", "Client comments"]

INK   = "1F3B57"
HEAD  = PatternFill("solid", fgColor=INK)
PEND  = PatternFill("solid", fgColor="FFF2CC")
ALT   = PatternFill("solid", fgColor="F4F7FA")
RCOL  = PatternFill("solid", fgColor="EAF3EA")
HFONT = Font(bold=True, color="FFFFFF", size=10)
THIN  = Side(style="thin", color="D6DEE6")
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP   = Alignment(vertical="top", wrap_text=True)

wb = Workbook()

def sheet(title, headers, widths, rows, notes=None, pending_col=None, freeze="A2"):
    ws = wb.create_sheet(title)
    r = 1
    if notes:
        for n in notes:
            c = ws.cell(row=r, column=1, value=n)
            c.font = Font(italic=True, size=9, color="6A7B8C")
            r += 1
        r += 1
    hrow = r
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=hrow, column=i, value=h)
        c.fill = HEAD; c.font = HFONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[hrow].height = 30
    nrev = len(REVIEW)
    first_rev = len(headers) - nrev + 1
    for j, row in enumerate(rows):
        rr = hrow + 1 + j
        for i, v in enumerate(row, 1):
            c = ws.cell(row=rr, column=i, value=v)
            c.alignment = TOP; c.border = BOX
            c.font = Font(size=10)
            if i >= first_rev:
                c.fill = RCOL
            elif pending_col is not None and row[pending_col]:
                c.fill = PEND
            elif j % 2:
                c.fill = ALT
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
    return ws

# ---------- 1. Read me ----------
ws = wb.active; ws.title = "Read me"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 112
def line(t, style=None, r=[1]):
    c = ws.cell(row=r[0], column=2, value=t)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if style == "h1": c.font = Font(bold=True, size=17, color=INK)
    elif style == "h2": c.font = Font(bold=True, size=11, color=INK)
    elif style == "flag": c.font = Font(size=10, color="9C6500")
    else: c.font = Font(size=10)
    ws.row_dimensions[r[0]].height = None if style else 14
    r[0] += 1
    return r[0]

line("Unlock AI Value: content for client review", "h1")
line(f"Insurance kiosk · exported {datetime.date.today():%d %B %Y} from the build's content model", "flag")
line("")
line("What this is", "h2")
line("Every word the kiosk displays, pulled straight out of the build so what you review is what visitors read. One sheet per content type.")
line("")
line("How to give us changes", "h2")
line("Three columns on the right of every sheet, shaded green:")
line("    Approved? (Y/N)      leave blank if you have not looked at the row yet")
line("    Revised copy         paste the wording you want instead. Full replacement, not a note")
line("    Client comments      anything that needs a conversation rather than a rewrite")
line("Do not edit the left-hand columns. We diff against them to find what moved.")
line("")
line("Where we still need content from Infosys", "h2")
line("These are gaps in the build today, not questions about wording. Highlighted amber on their sheets.")
line("    1.  Six of the eighteen proof tiles have no case study, marked 'Infosys to supply'. Physical AI has no case studies at all; AI Trust has one of three; Agentic Legacy Modernization has two of three. See 'Proof tiles'.")
line("        These six are now more urgent: one case study per pool also appears on that pool's diagnostic question, so Physical AI's question has no proof on it at all.")
line("    2.  Benchmark. Agreed the industry baseline comes from Gartner. Three things are needed and all three need Gartner access: the six baseline figures on a 0-100 scale, the exact reference (report title, author, publication date), and confirmation that Gartner's reprint terms allow attribution on a public kiosk. Until then the figures on screen stay labelled as placeholders and the fine print says so. See 'Benchmark'.")
line("    3.  The QR code on the final screen points nowhere yet. We need the destination URL.")
line("    5.  Case study clients are masked to descriptors, so no carrier is named anywhere in the build. If any are cleared for use by name, tell us which and we will restore them.")
line("    4.  The five-year forecast columns present today's case studies under future-dated headings. Flag if that framing is a problem.")
line("")
line("Sheet guide", "h2")
for t, d in [
    ("Diagnostic questions", "18 questions, 5 answers each (90 rows). Answer wording and the score behind each answer."),
    ("Value pools",          "The six pools: name, verb, description, three supporting facts."),
    ("Proof tiles",          "18 case-study tiles, 3 per pool. Six are unfilled, and client names are masked. One per pool also shows during the diagnostic."),
    ("Score band feedback",  "30 report blocks. What each score band tells a visitor, and the recommended move."),
    ("Archetypes",           "Five overall positions a visitor can land in."),
    ("Benchmark",            "Industry baseline per pool, to come from Gartner. Placeholder figures today."),
    ("Five-year view",       "One paragraph per archetype for the last results screen."),
    ("Screen copy",          "Headlines, buttons and instructional text that is not pool-specific."),
]:
    c = ws.cell(row=line.__defaults__[1][0], column=2)
    line(f"    {t:22}  {d}")

# ---------- 2. Diagnostic questions ----------
rows = []
for pid in ORDER:
    for qi, q in enumerate(D["QUESTIONS"][pid], 1):
        ref = f"{pid[:3].upper()}-{qi}"
        for oi, (lab, desc, score) in enumerate(q["opts"], 1):
            rows.append([
                NAME[pid], ref, q["kind"],
                q["q"] if oi == 1 else "",
                oi, lab, desc, score, "", "", ""
            ])
sheet("Diagnostic questions",
      ["Value pool", "Q ref", "Theme", "Question", "Stop",
       "Stop label (shown large)", "Stop detail (shown small)", "Score"] + REVIEW,
      [20, 8, 22, 54, 5, 26, 52, 6, 13, 42, 34], rows,
      notes=["Six questions, one per value pool, each answered on a slider that snaps to five stops. "
             "Was 18 questions with five tap-options each; Anshul asked for fewer of both.",
             "The five stops run from least to most mature, left to right, scoring 0 / 25 / 50 / 75 / 100. "
             "Those five scores are exactly the five report bands, so every stop has its own report copy "
             "on the 'Score band feedback' sheet.",
             "Stop labels are the large text on the slider, so keep them to about three or four words. "
             "The detail line below updates as the visitor drags."])

# ---------- 3. Value pools ----------
rows = [[NAME[p["id"]], p["verb"], " / ".join(p["lines"]), p["blurb"],
         p["facts"][0], p["facts"][1], p["facts"][2], "", "", ""]
        for p in sorted(D["POOLS"], key=lambda p: p["edge"])]
sheet("Value pools",
      ["Value pool", "Verb", "Board label", "Description", "Fact 1", "Fact 2", "Fact 3"] + REVIEW,
      [22, 14, 26, 60, 40, 40, 40, 13, 42, 34], rows,
      notes=["Shown when a visitor taps a pool on the hexagon. Platform names here are load-bearing, so flag any that are wrong or renamed."])

# ---------- 4. Proof tiles ----------
rows = []
for pid in ORDER:
    for t in D["TILES"][pid]:
        p = bool(t.get("pending"))
        rows.append([NAME[pid], "NEEDS CONTENT" if p else "Ready",
                     t.get("client", ""), t.get("title", ""), t.get("metric", ""),
                     t.get("detail", ""), "", "", "", p])
ws = sheet("Proof tiles",
      ["Value pool", "Status", "Client (masked)", "Tile title", "Headline metric", "Detail"] + REVIEW + ["_p"],
      [22, 15, 20, 28, 30, 66, 13, 42, 34, 3], rows, pending_col=9,
      notes=["Three per pool. Amber rows have no case study. We need client name, metric and one or two lines of detail, "
             "or permission to drop the tile.",
             "One of these appears in each diagnostic question panel, between the answer and Next, for that question's own value pool, "
             "per Anshul's note that proof should land while someone is still answering. The first cleared tile in "
             "each pool is used, and amber rows are skipped, so Physical AI currently shows no case study at all "
             "on its question.",
             "Client names are masked, per your note that all case study names go without client references. "
             "The Client column carries a descriptor instead. Tell us which carriers are cleared by name and we "
             "will put them back."])
ws.column_dimensions["J"].hidden = True

# ---------- 5. Score band feedback ----------
rows = []
for pid in ORDER:
    for bi, blk in enumerate(D["BAND_COPY"][pid]):
        rows.append([NAME[pid], BANDS[bi], RANGES[bi], blk["read"], blk["move"], "", "", ""])
sheet("Score band feedback",
      ["Value pool", "Band", "Score range", "What we tell them ('the read')", "What we recommend ('the move')"] + REVIEW,
      [22, 15, 11, 62, 72, 13, 42, 34], rows,
      notes=["The report body. A visitor sees one row per pool, whichever band their score falls into. 12 is the max per pool.",
             "This is where Infosys is making a recommendation to a prospect, so it is the highest-risk copy on the kiosk."])

# ---------- 6. Archetypes ----------
rows = [[a["name"], a["tag"], a["body"], a["risk"], "", "", ""]
        for a in D["ARCHETYPES"].values()]
sheet("Archetypes",
      ["Archetype", "Strapline", "Body", "Risk line"] + REVIEW,
      [24, 30, 76, 72, 13, 42, 34], rows,
      notes=["The visitor's overall position, chosen from their six pool scores. One of five."])

# ---------- 8. Benchmark ----------
rows = [[NAME[pid], D["BENCHMARK_MEDIAN"][pid], D["MAX_POOL_SCORE"], "", "", "", True]
        for pid in ORDER]
ws = sheet("Benchmark",
      ["Value pool", "Gartner baseline (placeholder today)", "Max score"] + REVIEW + ["_p"],
      [24, 26, 11, 13, 42, 34, 3], rows, pending_col=6,
      notes=[f"Fine print on the results screen and the final screen currently reads: \"{D['BENCHMARK_FINE_PRINT']}\"",
             "Every figure in this sheet is a placeholder we invented for layout. None of them is Gartner data and the "
             "screen does not claim otherwise.",
             "To go live we need three things from a Gartner seat: these six figures on the 0-100 scale, the exact "
             "reference to cite (report title, author, publication date), and confirmation that Gartner's reprint terms "
             "permit attribution on a public kiosk at a show. The third one is the long pole, so worth starting early.",
             "Once they arrive the fine print becomes: \"Industry baseline from Gartner research on AI adoption in "
             "insurance. Your score for each value pool is compared against that baseline.\" plus the citation."])
ws.column_dimensions["G"].hidden = True

# ---------- 9. Five-year paragraphs ----------
ARCH_NAME = {k: v["name"] for k, v in D["ARCHETYPES"].items()}
rows = [[ARCH_NAME.get(k, k), k, v, "", "", ""] for k, v in D["FIVE_YEAR"].items()]
sheet("Five-year view",
      ["Archetype", "Key", "Paragraph"] + REVIEW,
      [26, 14, 96, 13, 42, 34], rows,
      notes=["The last results screen. One paragraph per archetype, replacing the three-column "
             "compounding / holding / exposed grid, per the note asking for a single paragraph in a persona style.",
             "{lead} and {lag} are filled in live with that visitor's strongest and weakest value pool, so the "
             "paragraph names their own pools rather than describing their category. Both always appear and are "
             "always different pools. Keep both tokens if you rewrite a paragraph.",
             "These are about trajectory on purpose, not about position. 'Your position' on the first results screen "
             "already says where someone is and what the risk is, so repeating it here would be the third time the "
             "report makes the same point."])

# ---------- 10. Screen copy ----------
SCREEN = [
 ("Attract", "Eyebrow", "AI-First Value Framework"),
 ("Attract", "Headline", "Are you ready to unlock your AI Value in insurance?"),
 ("Attract", "Subhead", "Two minutes. Six value pools. One blueprint."),
 ("Attract", "Touch cue", "Touch to begin"),
 ("All screens", "Brand line", "Unlock AI Value in Insurance"),
 ("Framework screen", "Eyebrow", "An Infosys framework"),
 ("Framework screen", "Headline", "Six pools of AI value in insurance"),
 ("Framework screen", "Lede", "Infosys' own model for insurance, not an industry standard. Each pool "
                              "is somewhere AI creates measurable value for carriers, and somewhere "
                              "programmes commonly stall."),
 ("Framework screen", "Gist: AI Strategy & Engineering", "One operating model, instead of pilots scattered across the business"),
 ("Framework screen", "Gist: Data for AI", "Policy, claims and treaty data made AI-ready and audit-ready"),
 ("Framework screen", "Gist: Process AI", "Whole journeys redesigned, not task-level automation inside old ones"),
 ("Framework screen", "Gist: Agentic Legacy Modernization", "Core platforms modernised module by module, while the book stays live"),
 ("Framework screen", "Gist: Physical AI", "Telematics, IoT and drone data as a pricing and claims input"),
 ("Framework screen", "Gist: AI Trust", "Explainability, bias testing and audit trails built in from the start"),
 ("Framework screen", "Button (primary)", "Explore the six pools"),
 ("Framework screen", "Button (secondary)", "Start the diagnostic"),
 ("Explore", "Instruction", "Tap a value pool to explore"),
 ("Explore", "Button", "Start the diagnostic"),
 ("Diagnostic", "Progress", "Question 1 of 6"),
 ("Diagnostic", "Case card eyebrow", "[Value pool] \u00b7 [Client]"),
 ("Diagnostic", "Case card link", "Read the case"),
 ("Email capture", "Headline", "Your blueprint is ready"),
 ("Email capture", "Lede", "Add your email and we will send the full report, with the case studies behind every score."),
 ("Email capture", "Field label", "Work email"),
 ("Email capture", "Field placeholder", "name@company.com"),
 ("Email capture", "Button", "See my results"),
 ("Email capture", "Error", "Enter a valid email address."),
 ("Email capture", "Confirmation headline", "Your blueprint is on its way"),
 ("Email capture", "Confirmation body", "Sent to [email]. Your results are next."),
 ("Results 1", "Eyebrow", "Your position"),
 ("Results 2", "Headline", "Against the benchmark"),
 ("Results 3", "Eyebrow", "Five-year view"),
 ("Results 3", "Headline", "Where the gap goes"),
 ("Results 3", "Paragraph", "One per archetype, see the 'Five-year view' sheet"),
 ("Results 3", "Proof label", "Proven now \u00b7 [Value pool]"),
 ("Done", "Eyebrow", "Report dispatched"),
 ("Done", "Headline", "On its way"),
 ("Done", "Body", "Sent to [email]. Three ways to get it back:"),
 ("Done", "Step 1", "Scan the code above"),
 ("Done", "Step 2", "Open the email we just sent"),
 ("Done", "Step 3", "Scan your badge at the booth to see and discuss your result"),
]
rows = [[s, k, t, "", "", ""] for s, k, t in SCREEN]
sheet("Screen copy",
      ["Screen", "Element", "Text"] + REVIEW,
      [16, 16, 86, 13, 42, 34], rows,
      notes=["Fixed copy that is not pool- or score-specific. Square brackets are filled in live by the kiosk."])

wb.save(OUT)
print("saved:", OUT)
print("sheets:", ", ".join(wb.sheetnames))
